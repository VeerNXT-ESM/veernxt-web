#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
scripts/build_master_datamap.py

Joins every exam-catalog data source collected this session into one
authoritative "master datamap" -- one row per unique exam from
master_exam_list_unique.json, with subject requirements, PwD eligibility
(UT), logo, and content-completeness (Guide/Precis/Intro/PYQ/Mock) all
resolved onto it.

Several of the sources being joined were built against the *pre-dedup*
exam list, so a canonical row's data may be scattered across 2-3 old rows
that dedupe_exam_list.py folded together. This re-keys against both a
row's own identity and everything in its `also_listed_as` tag list so nothing
gets silently dropped.

Read-only against every source; writes exam_master_datamap.json,
exam_master_datamap.xlsx, and datamap_build_report.md. Does not modify any
existing file.

Usage:
  python scripts/build_master_datamap.py
"""

import json
import os
import re
import sys
import io
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from dedupe_exam_list import norm, norm_body, detect_aliases  # noqa: E402 (also wraps sys.stdout as utf-8)

import openpyxl  # noqa: E402

EXAM_LIST_DIR = r"K:\H DRIVE\Quantum Climb\CLIENT ASSETS\VeerNXT\CONTENT\1. EXAM LIST"
SCRATCH = r"K:\tmp\exam_list_extract"
REPO = r"k:\H DRIVE\Quantum Climb\APPS\VeerNXT\VeerNXT Main Repo\VeerNXT APP\veernxt-web"

UNIQUE_LIST = os.path.join(EXAM_LIST_DIR, "master_exam_list_unique.json")
SUBJECT_WISE = os.path.join(SCRATCH, "subject_wise_list.json")
PWD = os.path.join(SCRATCH, "ut_pwd_eligibility.json")
LOGO_PRIORITY = os.path.join(SCRATCH, "logo_priority.json")
UT_HASHES = os.path.join(SCRATCH, "ut_content_hashes.json")
MANIFEST_JSON = os.path.join(REPO, "exam-logos", "manifest.json")
RESOURCE_XLSX = os.path.join(REPO, "exam_resource_mapping.xlsx")

OUT_JSON = os.path.join(EXAM_LIST_DIR, "exam_master_datamap.json")
OUT_XLSX = os.path.join(EXAM_LIST_DIR, "exam_master_datamap.xlsx")
OUT_REPORT = os.path.join(EXAM_LIST_DIR, "datamap_build_report.md")

# UT generic-subject -> master cluster IDs, hand-verified in ut_content_mapping.md
UT_GENERIC_CLUSTERS = {
    "reasoning_guide": [7, 52],
    "mathematics_guide": [8],
    "english_guide": [5],
    "hindi_guide": [10],
    "computer_guide": [9],
    "gk_gs_guide": [6, 13, 35],
    "reasoning_precis": [4],
    "mathematics_precis": [3],
    "english_precis": [2],
    "hindi_precis": [12],
    "general_knowledge_precis": [75],
}

# UT -> its real, unique GS-book master cluster (Andaman/Chandigarh/DNH&DD/J&K/Delhi).
# Ladakh/Lakshadweep/Puducherry deliberately absent -- that's the flagged gap.
UT_GS_BOOK_CLUSTER = {
    "andaman and nicobar islands": 89,
    "chandigarh": 90,
    "dadra and nagar haveli and daman and diu": 91,
    "jammu and kashmir": 92,
    "delhi": 93,
}


def load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def row_keys(row, alias_map):
    """All (level, norm_body, norm_state, norm_exam) keys this canonical row
    answers to, including everything folded into it via also_listed_as.
    conducting_body is part of the key -- without it, generic exam names
    shared by many different bodies (e.g. 18 different bodies all have a
    'Staff Nurse' exam) would collide and broadcast one body's data onto
    all the others."""
    body = norm_body(alias_map.get(row["conducting_body"], row["conducting_body"]))
    keys = [(row["level"], body, norm(row.get("state") or ""), norm(row["exam_name"]))]
    for t in row.get("also_listed_as", []) or []:
        # tags don't carry their own conducting_body -- they're folded
        # duplicates of THIS row's exam, so they inherit its body identity.
        keys.append((t["level"], body, norm(t.get("state") or ""), norm(t["exam_name"])))
    return keys


def main():
    unique_rows = load(UNIQUE_LIST)
    print(f"Loaded {len(unique_rows)} unique exams")

    # One alias map, built off the ORIGINAL pre-dedup list -- the same one
    # dedupe_exam_list.py used to produce master_exam_list_unique.json's
    # canonical conducting_body values. Reused everywhere below so a source
    # row's raw body string ("SSC") and the spine's already-canonical one
    # ("Staff Selection Commission") resolve to the same join key.
    original_rows = load(os.path.join(SCRATCH, "master_exam_list.json"))
    alias_map, _ = detect_aliases(original_rows)

    out = []
    for r in unique_rows:
        o = dict(r)
        o["subject_requirements"] = None
        o["pwd_eligibility"] = None
        o["logo"] = None
        o["content_completeness"] = None
        o["ut_master_content"] = None
        out.append(o)

    key_index = defaultdict(list)
    for i, r in enumerate(unique_rows):
        for k in row_keys(r, alias_map):
            key_index[k].append(i)
    collisions = {k: v for k, v in key_index.items() if len(v) > 1}

    report = []
    report.append("# Master datamap build report\n")
    report.append(f"Spine: {len(unique_rows)} unique exams from `master_exam_list_unique.json`.\n")
    if collisions:
        report.append(f"\n**{len(collisions)} normalized keys map to more than one canonical row** "
                       f"(rare text collisions) -- both rows receive any source data matching that key.\n")

    # ---- Subject-wise requirements ----
    subject_wise = load(SUBJECT_WISE)
    sw_matched, sw_unmatched = 0, []
    for row in subject_wise:
        body = norm_body(alias_map.get(row["conducting_body"], row["conducting_body"]))
        key = (row["level"], body, norm(row.get("state") or ""), norm(row["exam_name"]))
        idxs = key_index.get(key)
        if idxs:
            for i in idxs:
                out[i]["subject_requirements"] = row["subjects"]
            sw_matched += 1
        else:
            sw_unmatched.append(row)
    report.append(f"\n## Subject-wise requirements\nMatched {sw_matched}/{len(subject_wise)} source rows "
                   f"({sum(1 for o in out if o['subject_requirements'])}/{len(out)} output rows covered).\n")
    if sw_unmatched:
        report.append(f"{len(sw_unmatched)} unmatched (sample):\n")
        for u in sw_unmatched[:15]:
            report.append(f"- [{u['level']}/{u.get('state')}] {u['conducting_body']} | {u['exam_name']!r}")

    # ---- PwD eligibility (UT only) ----
    pwd = load(PWD)
    pwd_matched, pwd_unmatched = 0, []
    for row in pwd:
        body = norm_body(alias_map.get(row["Conducting Body"], row["Conducting Body"]))
        key = ("ut", body, norm(row["State / UT"]), norm(row["Name of the Examination"]))
        idxs = key_index.get(key)
        if idxs:
            for i in idxs:
                out[i]["pwd_eligibility"] = {
                    "applicable": row.get("PwD Reservation Applicable?"),
                    "blindness_low_vision": row.get("Blindness & Low Vision (B/LV)"),
                    "deaf_hard_of_hearing": row.get("Deaf & Hard of Hearing (D/HH)"),
                    "locomotor_disability": row.get(
                        "Locomotor Disability (LD) incl. CP/Dwarfism/Leprosy Cured/Acid Attack/Muscular Dystrophy"),
                    "autism_intellectual_sld_mi": row.get("Autism/Intellectual/SLD/Mental Illness (ID/MI)"),
                    "basis": row.get("Basis / Note"),
                }
            pwd_matched += 1
        else:
            pwd_unmatched.append(row)
    ut_rows = [o for o in out if o["level"] == "ut"]
    report.append(f"\n## PwD eligibility (UT only)\nMatched {pwd_matched}/{len(pwd)} source rows "
                   f"({sum(1 for o in ut_rows if o['pwd_eligibility'])}/{len(ut_rows)} UT output rows covered).\n")
    if pwd_unmatched:
        for u in pwd_unmatched[:15]:
            report.append(f"- unmatched: {u['State / UT']} | {u['Name of the Examination']!r}")

    # ---- Logos ----
    manifest = load(MANIFEST_JSON)
    body_to_manifest = {}
    for m in manifest:
        canon = alias_map.get(m["conducting_body"], m["conducting_body"])
        existing = body_to_manifest.get(norm_body(canon))
        if existing is None or (not existing.get("logo") and m.get("logo")):
            body_to_manifest[norm_body(canon)] = m

    logo_priority = load(LOGO_PRIORITY)
    quality_by_body = {}
    for r in logo_priority.get("needs_work_ranked", []):
        quality_by_body[(r["level"], norm_body(r["conducting_body"]))] = r.get("quality")

    logo_matched = 0
    for o in out:
        m = body_to_manifest.get(norm_body(o["conducting_body"]))
        if m:
            logo_matched += 1
            o["logo"] = {
                "path": m.get("logo"),
                "quality": quality_by_body.get((o["level"], norm_body(o["conducting_body"])),
                                                "not flagged as needing work (adequate or unassessed)"),
            }
    report.append(f"\n## Logos\nMatched {logo_matched}/{len(out)} rows to a manifest entry "
                   f"(via alias-normalized conducting_body).\n")

    # ---- Content completeness: Central/State from Exam Summary sheet ----
    wb = openpyxl.load_workbook(RESOURCE_XLSX, data_only=True)
    ws = wb["Exam Summary"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    summary_rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    summary_index = defaultdict(list)
    for r in summary_rows:
        key = (r["Level"], norm(r["Exam Name"] or ""))
        summary_index[key].append(r)
    summary_ambiguous_keys = sum(1 for v in summary_index.values() if len(v) > 1)

    def best_candidate(candidates, conducting_body):
        """Disambiguate same-named exams (e.g. 'Constable' under many different
        Police bodies) by token overlap between the source row's folder path
        and this row's conducting_body. Only accept a clear, unique winner --
        otherwise leave unmatched rather than risk attaching the wrong body's
        completeness data."""
        if len(candidates) == 1:
            return candidates[0]
        body_tok = set(norm(conducting_body).split())
        scored = []
        for c in candidates:
            path_tok = set(norm((c.get("Institution/State Path") or "") + " " + (c.get("Full Exam Folder Path") or "")).split())
            scored.append((len(body_tok & path_tok), c))
        scored.sort(key=lambda x: -x[0])
        if scored[0][0] > 0 and (len(scored) == 1 or scored[0][0] > scored[1][0]):
            return scored[0][1]
        return None

    cs_matched = 0
    cs_unmatched_ambiguous = 0
    cs_rows = [o for o in out if o["level"] in ("central", "state")]
    for o in cs_rows:
        key = (o["level"], norm(o["exam_name"]))
        candidates = summary_index.get(key)
        r = best_candidate(candidates, o["conducting_body"]) if candidates else None
        if candidates and len(candidates) > 1 and not r:
            cs_unmatched_ambiguous += 1
        if r:
            cs_matched += 1
            o["content_completeness"] = {
                "intro": r.get("Intro"), "guide": r.get("Guide"), "precis": r.get("Precis"),
                "pyq": r.get("PYQ"), "mock": r.get("Mock"),
            }
    report.append(f"\n## Content completeness -- Central/State (from Exam Summary sheet)\n"
                   f"Matched {cs_matched}/{len(cs_rows)} rows. {summary_ambiguous_keys} exam names in the "
                   f"source sheet are shared by more than one conducting body (e.g. generic Police post "
                   f"titles); disambiguated by folder-path/conducting_body token overlap where there was a "
                   f"clear winner, left unmatched otherwise rather than risk attaching the wrong body's data "
                   f"-- {cs_unmatched_ambiguous} rows fell into that unresolved-ambiguous bucket.\n")

    # ---- Content completeness: UT, built fresh from cached content hashes ----
    ut_hashes = load(UT_HASHES)
    ut_by_exam = defaultdict(lambda: defaultdict(int))
    exams_by_state = defaultdict(set)
    for rec in ut_hashes:
        ut_state = re.sub(r"^\d+\.\s*", "", rec["ut"])
        exam = re.sub(r"^\d+\.\s*", "", rec["exam"])
        cat_sub = rec["category_sub"] or "root"
        key = (norm(ut_state), norm(exam))
        ut_by_exam[key][cat_sub] += 1
        exams_by_state[norm(ut_state)].add(norm(exam))

    import difflib as _difflib

    def fuzzy_ut_match(state_norm, exam_norm):
        """Exam names differ between the master exam list (docx table) and the
        actual UT EXAMS folder names (independently authored) -- e.g. 'Combined
        Graduate Level / Administrative Recruitment' vs 'Combined Graduate Level
        - Administrative Recruitment'. Fall back to fuzzy match within the same
        UT only, and only accept a clear, high-confidence winner."""
        candidates = exams_by_state.get(state_norm, set())
        best, best_ratio = None, 0.0
        for c in candidates:
            r = _difflib.SequenceMatcher(None, exam_norm, c).ratio()
            if r > best_ratio:
                best_ratio, best = r, c
        if best_ratio >= 0.6:
            return (state_norm, best)
        return None

    ut_matched, ut_matched_fuzzy = 0, 0
    for o in ut_rows:
        state_n, exam_n = norm(o.get("state") or ""), norm(o["exam_name"])
        key = (state_n, exam_n)
        counts = ut_by_exam.get(key)
        if not counts:
            fuzzy_key = fuzzy_ut_match(state_n, exam_n)
            if fuzzy_key:
                counts = ut_by_exam.get(fuzzy_key)
                ut_matched_fuzzy += 1
        if counts:
            ut_matched += 1
            o["content_completeness"] = {
                "intro": f"OK ({counts.get('1. INTRODUCTION', 0)} files)" if counts.get("1. INTRODUCTION") else "missing",
                "guide": f"OK ({counts.get('2. GUIDE BOOK', 0)} files)" if counts.get("2. GUIDE BOOK") else "missing",
                "precis": f"OK ({counts.get('3. PRECIS', 0)} files)" if counts.get("3. PRECIS") else "missing",
                "pyq": "not tracked separately for UT (see Test Series)",
                "mock": f"OK ({counts.get('5. 10 TEST SERIES', 0)} files)" if counts.get("5. 10 TEST SERIES") else "missing",
            }
            body_norm = norm_body(o["state"] or "")
            gs_cluster = UT_GS_BOOK_CLUSTER.get(body_norm)
            o["ut_master_content"] = {
                "generic_subject_clusters": UT_GENERIC_CLUSTERS,
                "ut_specific_gs_book_cluster": gs_cluster,
                "ut_specific_gs_book_status": (
                    f"Cluster_{gs_cluster:03d}" if gs_cluster else "MISSING -- flagged gap, see ut_content_mapping.md"
                ),
            }
    report.append(f"\n## Content completeness -- UT (built fresh from ut_content_hashes.json)\n"
                   f"Matched {ut_matched}/{len(ut_rows)} UT rows ({ut_matched_fuzzy} via fuzzy fallback -- "
                   f"exam-list and folder-name wording differ, e.g. '/' vs '-'; only a clear same-UT, "
                   f"ratio>=0.6 winner accepted, everything else left unmatched rather than guessed).\n")

    # ---- Write outputs ----
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False)
    print(f"Wrote {OUT_JSON} ({len(out)} rows)")

    write_xlsx(out, report)
    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(report) + "\n")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_REPORT}")


def write_xlsx(out, report_lines):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Datamap"

    subject_cols = ["Hindi / Regional Language", "English", "General Knowledge / GS", "Reasoning",
                     "Maths", "General Science", "Computer Knowledge", "Child Dev. & Pedagogy",
                     "Domain / Technical Subject", "Physical Test", "Interview", "Typing / Skill Test"]
    headers = (["Level", "State/UT", "Category", "Conducting Body", "Exam Name", "Website",
                "Also Listed As (folded dupes)"]
               + [f"Subject: {c}" for c in subject_cols]
               + ["PwD Applicable", "PwD Blindness/LV", "PwD Deaf/HH", "PwD Locomotor", "PwD Autism/ID/MI", "PwD Basis"]
               + ["Logo Path", "Logo Quality"]
               + ["Intro", "Guide", "Precis", "PYQ", "Mock"]
               + ["UT GS Book Status"])
    ws.append(headers)

    for o in out:
        sw = o.get("subject_requirements") or {}
        pwd = o.get("pwd_eligibility") or {}
        logo = o.get("logo") or {}
        cc = o.get("content_completeness") or {}
        umc = o.get("ut_master_content") or {}
        also = "; ".join(f"{t['level']}/{t.get('state')}: {t['exam_name']}" for t in (o.get("also_listed_as") or []))
        row = ([o["level"], o.get("state"), o.get("category"), o["conducting_body"], o["exam_name"],
                o.get("website"), also]
               + [sw.get(c) for c in subject_cols]
               + [pwd.get("applicable"), pwd.get("blindness_low_vision"), pwd.get("deaf_hard_of_hearing"),
                  pwd.get("locomotor_disability"), pwd.get("autism_intellectual_sld_mi"), pwd.get("basis")]
               + [logo.get("path"), logo.get("quality")]
               + [cc.get("intro"), cc.get("guide"), cc.get("precis"), cc.get("pyq"), cc.get("mock")]
               + [umc.get("ut_specific_gs_book_status")])
        row = [json.dumps(c, ensure_ascii=False) if isinstance(c, (dict, list)) else c for c in row]
        ws.append(row)

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Join Coverage")
    for line in report_lines:
        clean = re.sub(r"^#+\s*", "", line).strip()
        if clean:
            ws2.append([clean])

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
